"""
=============================================================================
FUZZY NAME MATCHER — Versión Streamlit (Mobile-friendly)
=============================================================================
Instrucciones para correr:
  pip install streamlit pandas rapidfuzz openpyxl
  streamlit run app.py

O en Streamlit Cloud:
  1. Sube este archivo a GitHub
  2. Ve a share.streamlit.io y conecta el repo
=============================================================================
"""

import streamlit as st
import pandas as pd
from rapidfuzz import fuzz, process
import unicodedata
import re
from datetime import datetime
import io

# ── Configuración de página ────────────────────────────────────────────────
st.set_page_config(
    page_title="Fuzzy Name Matcher",
    page_icon="🔍",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── CSS personalizado (mobile-first, diseño limpio) ───────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Fondo general */
.stApp {
    background: linear-gradient(135deg, #0f0c29, #302b63, #24243e);
    min-height: 100vh;
}

/* Título principal */
.titulo-principal {
    font-family: 'Space Mono', monospace;
    font-size: clamp(1.4rem, 5vw, 2.2rem);
    font-weight: 700;
    color: #ffffff;
    text-align: center;
    padding: 1.2rem 0 0.3rem;
    letter-spacing: -0.5px;
}

.subtitulo {
    text-align: center;
    color: #a78bfa;
    font-size: 0.95rem;
    margin-bottom: 2rem;
    font-weight: 300;
}

/* Tarjetas / secciones */
.card {
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 16px;
    padding: 1.2rem 1.4rem;
    margin-bottom: 1rem;
    backdrop-filter: blur(10px);
}

.card-title {
    font-family: 'Space Mono', monospace;
    font-size: 0.8rem;
    color: #a78bfa;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    margin-bottom: 0.8rem;
}

/* Métricas */
.metric-row {
    display: flex;
    gap: 0.8rem;
    justify-content: center;
    flex-wrap: wrap;
    margin: 1.2rem 0;
}

.metric-box {
    background: rgba(167,139,250,0.15);
    border: 1px solid rgba(167,139,250,0.3);
    border-radius: 12px;
    padding: 0.8rem 1.2rem;
    text-align: center;
    min-width: 100px;
    flex: 1;
}

.metric-num {
    font-family: 'Space Mono', monospace;
    font-size: 1.6rem;
    font-weight: 700;
    color: #fff;
    line-height: 1;
}

.metric-lbl {
    font-size: 0.72rem;
    color: #c4b5fd;
    margin-top: 0.3rem;
    text-transform: uppercase;
    letter-spacing: 0.8px;
}

/* Badges de similitud */
.badge-high   { background:#16a34a22; color:#4ade80; border:1px solid #4ade8055; padding:2px 10px; border-radius:20px; font-size:0.82rem; }
.badge-medium { background:#d9770622; color:#fb923c; border:1px solid #fb923c55; padding:2px 10px; border-radius:20px; font-size:0.82rem; }
.badge-low    { background:#dc262622; color:#f87171; border:1px solid #f8717155; padding:2px 10px; border-radius:20px; font-size:0.82rem; }
.badge-none   { background:#71717a22; color:#a1a1aa; border:1px solid #a1a1aa44; padding:2px 10px; border-radius:20px; font-size:0.82rem; }

/* Botón principal */
div.stButton > button {
    width: 100%;
    background: linear-gradient(135deg, #7c3aed, #a855f7);
    color: white;
    border: none;
    border-radius: 12px;
    padding: 0.75rem 1.5rem;
    font-family: 'Space Mono', monospace;
    font-size: 0.95rem;
    font-weight: 700;
    letter-spacing: 0.5px;
    transition: all 0.2s;
    cursor: pointer;
}
div.stButton > button:hover {
    background: linear-gradient(135deg, #6d28d9, #9333ea);
    transform: translateY(-1px);
    box-shadow: 0 8px 25px rgba(124,58,237,0.4);
}

/* Slider */
.stSlider > div { color: #fff !important; }
.stSlider [data-testid="stTickBar"] { color: #a78bfa !important; }

/* File uploader */
[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.04) !important;
    border: 1.5px dashed rgba(167,139,250,0.4) !important;
    border-radius: 12px !important;
}

/* Select box */
.stSelectbox > div > div {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(255,255,255,0.15) !important;
    border-radius: 10px !important;
    color: white !important;
}

/* Labels */
label, .stSelectbox label, .stSlider label, .stFileUploader label {
    color: #e2e8f0 !important;
    font-size: 0.9rem !important;
    font-weight: 600 !important;
}

/* Download button */
[data-testid="stDownloadButton"] > button {
    width: 100%;
    background: linear-gradient(135deg, #059669, #10b981) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    font-family: 'Space Mono', monospace !important;
    font-weight: 700 !important;
}

/* Tabla de resultados */
.result-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0.7rem 0;
    border-bottom: 1px solid rgba(255,255,255,0.07);
    gap: 0.5rem;
    flex-wrap: wrap;
}
.result-name { color: #f1f5f9; font-size: 0.88rem; flex: 1; min-width: 120px; }
.result-match { color: #a78bfa; font-size: 0.85rem; flex: 1; min-width: 120px; }
.result-pct { font-family: 'Space Mono', monospace; font-size: 0.9rem; font-weight:700; min-width: 50px; text-align:right; }

/* Divider */
hr { border-color: rgba(255,255,255,0.08) !important; }

/* Info/warning/success boxes */
.stAlert { border-radius: 12px !important; }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# FUNCIONES DE NORMALIZACIÓN Y COMPARACIÓN
# =============================================================================

def normalize_name(name: str) -> str:
    if not isinstance(name, str):
        return ""
    name = name.lower().strip()
    name = unicodedata.normalize("NFD", name)
    name = "".join(c for c in name if unicodedata.category(c) != "Mn")
    name = re.sub(r"[^a-z\s]", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def tokenize(name: str) -> set:
    return set(normalize_name(name).split())


def hybrid_similarity(name1: str, name2: str) -> float:
    n1 = normalize_name(name1)
    n2 = normalize_name(name2)
    if not n1 or not n2:
        return 0.0

    score_token_set  = fuzz.token_set_ratio(n1, n2)
    score_token_sort = fuzz.token_sort_ratio(n1, n2)
    score_partial    = fuzz.partial_ratio(n1, n2)

    base_score = (
        score_token_set  * 0.50 +
        score_token_sort * 0.30 +
        score_partial    * 0.20
    )

    tokens1 = tokenize(name1)
    tokens2 = tokenize(name2)
    if not tokens1 or not tokens2:
        return base_score

    common_tokens = tokens1 & tokens2
    shorter_len   = min(len(tokens1), len(tokens2))
    common_ratio  = len(common_tokens) / shorter_len if shorter_len > 0 else 0

    if common_ratio < 0.5:
        base_score *= common_ratio * 1.5

    shorter = tokens1 if len(tokens1) <= len(tokens2) else tokens2
    longer  = tokens1 if len(tokens1) >  len(tokens2) else tokens2
    if shorter.issubset(longer):
        base_score = min(100, base_score + 10)

    extra_words = len(tokens1 - tokens2) + len(tokens2 - tokens1)
    penalty     = extra_words * 8
    if base_score > 90:
        penalty *= 0.3
    elif base_score > 80:
        penalty *= 0.6

    return round(max(0.0, base_score - penalty), 2)


def find_best_match(name: str, candidates: list, min_score: float):
    if not candidates or not name:
        return None, 0.0
    norm_name = normalize_name(name)
    norm_cands = [normalize_name(c) for c in candidates]

    rapid_results = process.extract(
        norm_name, norm_cands,
        scorer=fuzz.token_set_ratio,
        limit=10, score_cutoff=40
    )
    if not rapid_results:
        return None, 0.0

    best_name, best_score = None, 0.0
    for _, _, idx in rapid_results:
        score = hybrid_similarity(name, candidates[idx])
        if score > best_score:
            best_score = score
            best_name  = candidates[idx]

    return (best_name, best_score) if best_score >= min_score else (None, 0.0)


def run_matching(df1, df2, col1, col2, min_sim, progress_bar, status_text):
    names1 = df1[col1].dropna().astype(str).tolist()
    names2 = df2[col2].dropna().astype(str).tolist()
    total  = len(names1)
    results = []

    for i, name in enumerate(names1):
        best_match, score = find_best_match(name, names2, min_sim)
        if best_match:
            estado = "Alta" if score >= 90 else "Media" if score >= 75 else "Baja"
        else:
            best_match = "— Sin coincidencia —"
            score      = 0.0
            estado     = "Sin match"

        results.append({
            "Nombre Tabla 1":              name,
            "Mejor Coincidencia Tabla 2":  best_match,
            "Similitud (%)":               score,
            "Estado":                      estado,
        })

        pct = (i + 1) / total
        progress_bar.progress(pct)
        status_text.markdown(
            f'<p style="color:#a78bfa;font-size:0.85rem;text-align:center;">'
            f'Procesando {i+1} de {total} nombres...</p>',
            unsafe_allow_html=True
        )

    return pd.DataFrame(results)


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Convierte el DataFrame a bytes de Excel con formato."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", fgColor="302b63")
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    bs          = Side(style="thin", color="CCCCCC")
    border      = Border(left=bs, right=bs, top=bs, bottom=bs)

    headers    = ["Nombre Tabla 1", "Mejor Coincidencia Tabla 2", "Similitud (%)", "Estado"]
    col_widths = [35, 35, 16, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    color_map = {"Alta": "C8F7C5", "Media": "FFF3CD", "Baja": "FDE8E8", "Sin match": "E8E8E8"}

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        score  = row["Similitud (%)"]
        estado = row["Estado"]
        fill   = PatternFill("solid", fgColor=color_map.get(estado, "FFFFFF"))
        font   = Font(name="Arial", size=10)

        vals   = [row["Nombre Tabla 1"], row["Mejor Coincidencia Tabla 2"],
                  f"{score:.1f}%" if score > 0 else "—", estado]
        aligns = [left, left, center, center]

        for ci, (v, a) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = fill; c.font = font; c.alignment = a; c.border = border

    # Hoja resumen
    ws2   = wb.create_sheet("Resumen")
    total = len(df)
    mat   = len(df[df["Similitud (%)"] > 0])

    for r, (lbl, val) in enumerate([
        ("Total procesados", total),
        ("Con coincidencia", mat),
        ("Sin coincidencia", total - mat),
        ("Alta similitud (≥90%)", len(df[df["Similitud (%)"] >= 90])),
        ("Media similitud (75-89%)", len(df[(df["Similitud (%)"] >= 75) & (df["Similitud (%)"] < 90)])),
        ("Baja similitud (<75%)",    len(df[(df["Similitud (%)"] > 0)  & (df["Similitud (%)"] < 75)])),
        ("Fecha", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ], 1):
        ws2.cell(row=r, column=1, value=lbl).font  = Font(bold=True, name="Arial")
        ws2.cell(row=r, column=2, value=str(val)).font = Font(name="Arial")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# INTERFAZ STREAMLIT
# =============================================================================

# Título
st.markdown('<div class="titulo-principal">🔍 Fuzzy Name Matcher</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitulo">Comparación inteligente de nombres entre archivos Excel</div>', unsafe_allow_html=True)

# ── PASO 1: Subir archivos ─────────────────────────────────────────────────
st.markdown('<div class="card"><div class="card-title">📂 Paso 1 — Cargar archivos</div>', unsafe_allow_html=True)

col_a, col_b = st.columns(2)
with col_a:
    file1 = st.file_uploader("Tabla 1 (.xlsx)", type=["xlsx", "xls"], key="f1")
with col_b:
    file2 = st.file_uploader("Tabla 2 (.xlsx)", type=["xlsx", "xls"], key="f2")

st.markdown('</div>', unsafe_allow_html=True)

df1, df2 = None, None

if file1:
    try:
        df1 = pd.read_excel(file1, dtype=str)
        st.success(f"✅ Tabla 1: **{file1.name}** — {len(df1):,} filas, {len(df1.columns)} columnas")
    except Exception as e:
        st.error(f"Error al leer Tabla 1: {e}")

if file2:
    try:
        df2 = pd.read_excel(file2, dtype=str)
        st.success(f"✅ Tabla 2: **{file2.name}** — {len(df2):,} filas, {len(df2.columns)} columnas")
    except Exception as e:
        st.error(f"Error al leer Tabla 2: {e}")

# ── PASO 2: Elegir columnas ────────────────────────────────────────────────
if df1 is not None and df2 is not None:

    st.markdown('<div class="card"><div class="card-title">📋 Paso 2 — Seleccionar columnas de nombres</div>', unsafe_allow_html=True)

    col_c, col_d = st.columns(2)
    with col_c:
        col1_sel = st.selectbox("Columna de nombres (Tabla 1)", df1.columns.tolist(), key="c1")
    with col_d:
        col2_sel = st.selectbox("Columna de nombres (Tabla 2)", df2.columns.tolist(), key="c2")

    st.markdown('</div>', unsafe_allow_html=True)

    # ── PASO 3: Configuración ──────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">⚙️ Paso 3 — Configuración</div>', unsafe_allow_html=True)

    min_sim = st.slider(
        "Similitud mínima requerida (%)",
        min_value=30, max_value=100, value=75, step=1,
        help="70-80% recomendado. Más alto = menos coincidencias pero más precisas."
    )

    st.markdown(
        f'<p style="color:#a78bfa; font-size:0.82rem; margin-top:-0.5rem;">'
        f'{"🟢 Estricto — pocos falsos positivos" if min_sim >= 85 else "🟡 Balanceado — recomendado" if min_sim >= 70 else "🔴 Permisivo — puede haber falsos positivos"}'
        f'</p>',
        unsafe_allow_html=True
    )
    st.markdown('</div>', unsafe_allow_html=True)

    # ── PASO 4: Ejecutar ───────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">▶️ Paso 4 — Ejecutar comparación</div>', unsafe_allow_html=True)

    n1 = len(df1[col1_sel].dropna())
    n2 = len(df2[col2_sel].dropna())
    st.markdown(
        f'<p style="color:#94a3b8; font-size:0.85rem;">'
        f'Se compararán <b style="color:#fff">{n1:,}</b> nombres de Tabla 1 contra '
        f'<b style="color:#fff">{n2:,}</b> nombres de Tabla 2.</p>',
        unsafe_allow_html=True
    )

    run_btn = st.button("🔍 Iniciar Comparación Inteligente")
    st.markdown('</div>', unsafe_allow_html=True)

    # ── PROCESO ────────────────────────────────────────────────────────────
    if run_btn:
        progress_bar = st.progress(0)
        status_text  = st.empty()

        with st.spinner(""):
            results_df = run_matching(df1, df2, col1_sel, col2_sel, min_sim, progress_bar, status_text)

        st.session_state["results_df"] = results_df
        status_text.empty()
        progress_bar.empty()

# ── RESULTADOS ─────────────────────────────────────────────────────────────
if "results_df" in st.session_state:
    df_res = st.session_state["results_df"]

    total    = len(df_res)
    matched  = len(df_res[df_res["Similitud (%)"] > 0])
    high     = len(df_res[df_res["Similitud (%)"] >= 90])
    medium   = len(df_res[(df_res["Similitud (%)"] >= 75) & (df_res["Similitud (%)"] < 90)])
    no_match = total - matched

    # Métricas
    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-box">
            <div class="metric-num">{total:,}</div>
            <div class="metric-lbl">Total</div>
        </div>
        <div class="metric-box">
            <div class="metric-num" style="color:#4ade80">{matched:,}</div>
            <div class="metric-lbl">Con match</div>
        </div>
        <div class="metric-box">
            <div class="metric-num" style="color:#f87171">{no_match:,}</div>
            <div class="metric-lbl">Sin match</div>
        </div>
        <div class="metric-box">
            <div class="metric-num" style="color:#a78bfa">{high:,}</div>
            <div class="metric-lbl">Alta sim.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Filtros
    st.markdown('<div class="card"><div class="card-title">🔎 Filtrar resultados</div>', unsafe_allow_html=True)
    filtro = st.selectbox("Mostrar", ["Todos", "Solo coincidencias", "Alta similitud (≥90%)", "Media (75-89%)", "Sin coincidencia"], key="filtro")
    st.markdown('</div>', unsafe_allow_html=True)

    df_show = df_res.copy()
    if filtro == "Solo coincidencias":
        df_show = df_show[df_show["Similitud (%)"] > 0]
    elif filtro == "Alta similitud (≥90%)":
        df_show = df_show[df_show["Similitud (%)"] >= 90]
    elif filtro == "Media (75-89%)":
        df_show = df_show[(df_show["Similitud (%)"] >= 75) & (df_show["Similitud (%)"] < 90)]
    elif filtro == "Sin coincidencia":
        df_show = df_show[df_show["Similitud (%)"] == 0]

    # Tabla de resultados
    st.markdown(f'<div class="card"><div class="card-title">📊 Resultados — {len(df_show):,} registros</div>', unsafe_allow_html=True)

    rows_html = ""
    for _, row in df_show.head(200).iterrows():
        score  = row["Similitud (%)"]
        estado = row["Estado"]
        pct_str = f"{score:.1f}%" if score > 0 else "—"

        if estado == "Alta":
            badge = f'<span class="badge-high">✅ {pct_str}</span>'
        elif estado == "Media":
            badge = f'<span class="badge-medium">⚠️ {pct_str}</span>'
        elif estado == "Baja":
            badge = f'<span class="badge-low">🔴 {pct_str}</span>'
        else:
            badge = f'<span class="badge-none">— Sin match</span>'

        rows_html += f"""
        <div class="result-row">
            <div class="result-name">📄 {row['Nombre Tabla 1']}</div>
            <div class="result-match">🔗 {row['Mejor Coincidencia Tabla 2']}</div>
            <div>{badge}</div>
        </div>"""

    if len(df_show) > 200:
        rows_html += f'<p style="color:#94a3b8; font-size:0.8rem; text-align:center; padding:0.5rem;">... y {len(df_show)-200} más. Descarga el Excel para ver todos.</p>'

    st.markdown(rows_html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Exportar
    st.markdown('<div class="card"><div class="card-title">💾 Descargar resultados</div>', unsafe_allow_html=True)
    excel_bytes = df_to_excel_bytes(df_res)
    fname = f"fuzzy_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="⬇️  Descargar Excel con todos los resultados",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.markdown('</div>', unsafe_allow_html=True)

elif df1 is None or df2 is None:
    st.markdown("""
    <div style="text-align:center; color:#64748b; padding:2rem 1rem;">
        <div style="font-size:3rem; margin-bottom:0.5rem;">📋</div>
        <div style="font-size:0.95rem;">Carga tus dos archivos Excel para comenzar</div>
    </div>
    """, unsafe_allow_html=True)
